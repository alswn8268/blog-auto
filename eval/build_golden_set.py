"""골든셋 구축 (명세 3장).

직원이 이미 만들어 운영 중인 Q&A 데이터가 있다면 처음부터 새로 만들 필요 없이
이걸 변환해서 쓴다. JSON/CSV/XLSX를 모두 받는다.

다만 그대로 쓰기 전에 두 가지는 꼭 확인할 것.
  - 지금도 유효한 답변인가: 그 사이 조항이 개정됐다면 답변도 같이 갱신해야 한다.
  - 질문 표현이 실제 사용자 말투와 비슷한가: 운영 Q&A는 보통 정제된 문어체라,
    실제 채팅창에 들어올 법한 구어체 질문 몇 개는 별도로 추가하는 것이 좋다.

실행: python -m eval.build_golden_set --input data/qa.xlsx --out eval/golden_set.json
"""

from __future__ import annotations

import argparse
import csv
import json
import logging
from pathlib import Path

logger = logging.getLogger(__name__)

# 실제 파일의 열 이름이 다르면 여기만 고치면 된다
QUESTION_KEYS = ("질문", "question", "Q", "문의내용")
ANSWER_KEYS = ("답변", "answer", "A", "회신내용")
SOURCE_KEYS = ("근거", "근거조항", "source", "출처")


def _pick(row: dict, keys: tuple[str, ...]) -> str:
    for k in keys:
        v = row.get(k)
        if v not in (None, ""):
            return str(v).strip()
    return ""


def _rows_from_file(filepath: Path) -> list[dict]:
    suffix = filepath.suffix.lower()
    if suffix == ".json":
        data = json.loads(filepath.read_text(encoding="utf-8"))
        return data if isinstance(data, list) else data.get("items", [])
    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with open(filepath, encoding="utf-8-sig", newline="") as f:
            return list(csv.DictReader(f, delimiter=delimiter))
    if suffix in {".xlsx", ".xlsm"}:
        from openpyxl import load_workbook  # 지연 임포트

        ws = load_workbook(filepath, read_only=True, data_only=True).active
        rows = ws.iter_rows(values_only=True)
        header = [str(h).strip() if h is not None else "" for h in next(rows)]
        return [dict(zip(header, r)) for r in rows]
    raise ValueError(f"지원하지 않는 형식입니다: {suffix}")


def load_existing_qa(filepath: str | Path) -> list[dict]:
    """직원이 만든 운영 Q&A 데이터를 골든셋 포맷으로 변환한다."""
    rows = _rows_from_file(Path(filepath))
    golden = []
    for row in rows:
        question, answer = _pick(row, QUESTION_KEYS), _pick(row, ANSWER_KEYS)
        if not question or not answer:
            continue
        item = {"question": question, "ground_truth": answer, "origin": "운영Q&A"}
        source = _pick(row, SOURCE_KEYS)
        if source:
            item["reference"] = source
        golden.append(item)

    if not golden:
        logger.warning("변환된 항목이 없습니다. 열 이름(QUESTION_KEYS/ANSWER_KEYS)을 확인하세요.")
    return golden


def merge_colloquial(golden: list[dict], extra_path: str | Path | None) -> list[dict]:
    """직접 추가한 구어체 질문을 합친다.

    운영 Q&A만으로는 실제 채팅 말투를 대표하지 못하므로, 구어체 셋을 반드시 섞는다.
    """
    if not extra_path:
        return golden
    extra = json.loads(Path(extra_path).read_text(encoding="utf-8"))
    for item in extra:
        item.setdefault("origin", "구어체추가")
    return golden + extra


def save_golden_set(items: list[dict], out_path: str | Path) -> Path:
    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(items, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(message)s")
    parser = argparse.ArgumentParser(description="운영 Q&A를 RAGAS 골든셋으로 변환한다")
    parser.add_argument("--input", required=True, help="운영 Q&A 파일 (json/csv/xlsx)")
    parser.add_argument("--colloquial", default=None, help="직접 추가한 구어체 질문 JSON")
    parser.add_argument("--out", default="eval/golden_set.json")
    args = parser.parse_args()

    golden = merge_colloquial(load_existing_qa(args.input), args.colloquial)
    out = save_golden_set(golden, args.out)
    logger.info("골든셋 %d문항 → %s", len(golden), out)
    logger.info("주의: 조항이 개정됐다면 ground_truth도 함께 최신화하세요.")


if __name__ == "__main__":
    main()
